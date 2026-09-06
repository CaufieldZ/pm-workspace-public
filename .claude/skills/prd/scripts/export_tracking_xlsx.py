#!/usr/bin/env python3
"""导出 PRD 埋点章节为 Excel（事件级列同事件合并单元格）。

读任意 PRD md（baseline §9.1 / delta §7 / community tracking md）里的 10 列埋点表，
按事件英文名归组，事件标识列（所属页面 / 事件中文名 / 事件英文名 / 应埋点平台）
做 vertical merge，属性级列每属性一行保留细节。供数据团队 / 研发对齐 + 粘贴
Confluence——md 表格不支持 rowspan，这里补上。

用法：python3 export_tracking_xlsx.py <prd.md> [-o out.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["所属页面", "事件中文名", "事件英文名", "属性中文名", "属性英文名",
           "数据类型", "是否必填", "数据值示例", "应埋点平台", "触发机制"]
# 事件级列（同事件合并）0 基索引：所属页面 / 事件中文名 / 事件英文名 / 应埋点平台
EVENT_LEVEL_COLS = {0, 1, 2, 8}
# 去反引号的英文标识符列：事件英文名 / 属性英文名
BACKTICK_COLS = {2, 4}


def _cells(line: str) -> list[str]:
    return [c.strip() for c in line.strip().strip("|").split("|")]


def _is_header(cells: list[str]) -> bool:
    """锚定列签名识别埋点表头（容「必填/是否必填」等措辞差异）。"""
    return len(cells) == 10 and cells[0] == "所属页面" and cells[2] == "事件英文名"


def parse_table(md_text: str) -> list[list[str]]:
    """按 10 列表头签名定位埋点表 → [[10 列], ...]，不依赖章节号。"""
    lines = md_text.splitlines()
    rows: list[list[str]] = []
    in_table = False
    for line in lines:
        s = line.strip()
        if not in_table:
            if s.startswith("|") and _is_header(_cells(s)):
                in_table = True
            continue
        if not s.startswith("|"):  # 表格结束（同签名多张表都收：重置状态，后续表头可再次进入）
            in_table = False
            continue
        cells = _cells(s)
        # 分隔行 = 每格都只由 : - 空格组成（含空格）；只看 cells[0] 会误删
        # 所属页面留空（合并单元格续行）或填 `-` 占位的真实埋点行
        if all(set(c) <= set(": -") for c in cells):
            continue
        if len(cells) < 10:
            continue
        for ci in BACKTICK_COLS:
            cells[ci] = cells[ci].strip("`")
        rows.append(cells[:10])
    return rows


def build_workbook(rows: list[list[str]]) -> tuple[Workbook, int]:
    """按事件英文名归组，事件级列 vertical merge。返回 (wb, 事件数)。"""
    groups: dict[str, list[list[str]]] = {}
    order: list[str] = []
    for r in rows:
        ev = r[2]
        if ev not in groups:
            groups[ev] = []
            order.append(ev)
        groups[ev].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "埋点清单"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.append(HEADERS)
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center, border

    row_idx = 2
    for ev in order:
        grp = groups[ev]
        start = row_idx
        for r in grp:
            ws.append(r)
            row_idx += 1
        end = row_idx - 1
        for ci in EVENT_LEVEL_COLS:
            if end > start:
                col = get_column_letter(ci + 1)
                ws.merge_cells(f"{col}{start}:{col}{end}")
            ws.cell(row=start, column=ci + 1).alignment = center
        for rr in range(start, end + 1):
            for ci in range(len(HEADERS)):
                cell = ws.cell(row=rr, column=ci + 1)
                cell.border = border
                if ci not in EVENT_LEVEL_COLS:
                    cell.alignment = wrap_top

    widths = [12, 22, 36, 20, 26, 10, 8, 34, 12, 42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return wb, len(order)


def main() -> int:
    ap = argparse.ArgumentParser(description="导出 PRD 埋点章节为合并单元格 xlsx")
    ap.add_argument("prd_md", help="PRD md 路径（baseline / delta / tracking md 均可）")
    ap.add_argument("-o", "--out", help="输出 xlsx 路径（默认 <prd 基名>-埋点.xlsx 落同目录）")
    args = ap.parse_args()

    src = Path(args.prd_md)
    if not src.is_file():
        print(f"找不到文件：{src}", file=sys.stderr)
        return 1
    out = Path(args.out) if args.out else src.with_name(f"{src.stem}-埋点.xlsx")

    rows = parse_table(src.read_text(encoding="utf-8", errors="replace"))
    if not rows:
        print(f"未在 {src} 找到 10 列埋点表（表头需为：{' / '.join(HEADERS)}）",
              file=sys.stderr)
        return 2

    wb, n_events = build_workbook(rows)
    wb.save(out)
    print(f"导出 {len(rows)} 属性行 / {n_events} 事件 → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
